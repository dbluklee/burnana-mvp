from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
import datetime
import time
import pandas as pd
import os

# --- 변수 설정 ---
max_iterations = 50  # 최대 반복횟수가 0이면 끝날때까지 계속 검색

# 경로 및 파일명 설정
current_path = os.path.dirname(os.path.abspath(__file__))

scraping_file_path = os.path.join(current_path, "scaping_data")
os.makedirs(scraping_file_path, exist_ok=True)

csv_file_name = os.path.join(current_path, "store_list.csv")


# 네이버 CSS 선택자
css_store_name = 'span.GHAhO'
css_store_category = 'span.lnJFt'
css_store_desc = 'div.XtBbS'
css_store_addr = 'span.LDgIH'
css_store_dir = 'span.zPfVt'
css_store_phone = 'span.xlx7Q'
css_store_phone_popup = 'a.BfF3H'
css_store_phone_popup_no = 'div.J7eF_ em'
css_store_sns = 'div.jO09N a'
css_store_etc = 'div.xPvPE'
css_store_intro = 'div.T8RFa'
css_store_services = 'li.c7TR6'
css_store_service = 'div.owG4q'


# --- CSV 파일에서 가게 ID 목록 읽어오기 ---
if not os.path.exists(csv_file_name):
    print(f"오류: '{csv_file_name}' 파일을 찾을 수 없습니다. 가게 ID가 포함된 CSV 파일을 생성해주세요.")
    exit() # 파일이 없으면 스크립트 종료

df = pd.read_csv(csv_file_name)
# 'id' 컬럼의 값을 리스트로 변환 (컬럼명이 다른 경우 이 부분을 수정하세요)
store_ids = df['id'].tolist() 
print(f"총 {len(store_ids)}개의 가게에 대한 리뷰 수집을 시작합니다.")


# WebDriver와 Workbook 리소스를 안전하게 관리하기 위해 try...finally 블록 안에서 초기화
driver = None
xlsx = None

try:
    # --- WebDriver 설정 ---
    options = Options()
    options.add_argument("window-size=1920x1080")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Safari/537.36") # User-Agent 추가
    service = Service(executable_path=ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)

    for store_id in store_ids:
        
        # --- 가게 기본 정보 로드 ---
        store_page = 'https://m.place.naver.com/restaurant/' + str(store_id) + '/home'
        driver.get(store_page)
        time.sleep(3) # 페이지 초기 로드를 위한 짧은 대기
        print("매장 기본정보 수집을 시작합니다...")

        # --- 엑셀 파일 및 시트 설정 ---
        xlsx = Workbook()

        # 📌 1. 첫 번째 시트: 'info' (가게 정보)
        info_sheet = xlsx.active # 기본으로 생성된 시트를 'info' 시트로 사용합니다.
        info_sheet.title = 'info'

        # 가게 정보 초기화
        store_name, store_category, store_desc = "", "", ""
        store_addr, store_dir, store_phone = "", "", ""
        store_sns, store_etc = "", ""
        store_intro, store_service = "", ""


        # --- 가게 정보 데이터 스크래핑 ---
        store_names = driver.find_elements(By.CSS_SELECTOR, css_store_name)
        if store_names:
            store_name = store_names[0].text.strip()
        info_sheet.append(['매장 이름', store_name])


        store_categorys = driver.find_elements(By.CSS_SELECTOR, css_store_category)
        if store_categorys:
            store_category = store_categorys[0].text.strip()
        info_sheet.append(['매장 유형', store_category])

        store_descs = driver.find_elements(By.CSS_SELECTOR, css_store_desc)
        if store_descs:
            store_desc = store_descs[0].text.strip()
        info_sheet.append(['매장 간단 소개', store_desc])

        store_addrs = driver.find_elements(By.CSS_SELECTOR, css_store_addr)
        if store_addrs:
            store_addr = store_addrs[0].text.strip()
        info_sheet.append(['매장 주소', store_addr])

        store_dirs = driver.find_elements(By.CSS_SELECTOR, css_store_dir)
        if store_dirs:
            store_dir = store_dirs[0].text.strip()
        info_sheet.append(['매장 오시는 길', store_dir])

        store_phones = driver.find_elements(By.CSS_SELECTOR, css_store_phone)
        if store_phones:
            store_phone = store_phones[0].text.strip()
        elif driver.find_elements(By.CSS_SELECTOR, css_store_phone_popup):
            try:
                phone_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, css_store_phone_popup))
                )
                # JavaScript를 사용하여 클릭 (클릭 오류 방지)
                driver.execute_script("arguments[0].click();", phone_button)
            except Exception as e:
                print(f"전화번호 버튼을 찾는 중 오류가 발생했습니다: {e}")
            store_phone = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, css_store_phone_popup_no))
                ).text.strip()
        else:
            store_phone = ''
        info_sheet.append(['매장 전화번호', store_phone])

        store_snss = driver.find_elements(By.CSS_SELECTOR, css_store_sns)
        if store_snss:
            store_sns = store_snss[0].get_attribute('href')
        info_sheet.append(['매장 sns', store_sns])

        store_etcs = driver.find_elements(By.CSS_SELECTOR, css_store_etc)
        if store_etcs:
            store_etc = store_etcs[0].text.strip()
        info_sheet.append(['기타 매장 정보', store_etc])


        # 가게 추가정보 페이지 로드
        store_page = 'https://m.place.naver.com/restaurant/' + str(store_id) + '/information'
        driver.get(store_page)
        time.sleep(3) # 페이지 초기 로드를 위한 짧은 대기
        print("매장 추가정보 수집을 시작합니다...")

        store_intros = driver.find_elements(By.CSS_SELECTOR, css_store_intro)
        if store_intros:
            store_intro = store_intros[0].text.strip()
        info_sheet.append(['매장 소개', store_intro])

        store_service = []
        store_services_list = driver.find_elements(By.CSS_SELECTOR, css_store_services)
        if store_services_list:
            for li in store_services_list:
                try:
                    service = li.find_element(By.CSS_SELECTOR, css_store_service).text.strip()
                except:
                    service = ''
            
                extra_services_list = li.find_elements(By.CSS_SELECTOR, 'span.place_blind')
                if extra_services_list:
                    extra_service = extra_services_list[0].text.strip()
                    service_name = f'{service} ({extra_service})'
                else:
                    service_name = service

                if service_name:
                    store_service.append(service_name)
        service_string = ", ".join(store_service)
        info_sheet.append(['매장 편의시설 및 서비스', service_string])

        # 가게 리뷰 페이지 로드
        store_page = 'https://m.place.naver.com/restaurant/' + str(store_id) + '/menu/list'
        driver.get(store_page)
        time.sleep(3) # 페이지 초기 로드를 위한 짧은 대기
 
        # 📌 2. 두 번째 시트: 'menu' (가게 후기)
        menu_sheet = xlsx.create_sheet('menu') # 'review'라는 이름의 새 시트를 생성합니다.
        menu_sheet.append(['메뉴명', '가격', '설명', '추천여부'])

        # --- '더보기' 버튼 클릭 반복 ---
        while True:
            try:
                # '더보기' 버튼이 클릭 가능할 때까지 최대 10초 대기
                more_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.fvwqf'))
                )
                # JavaScript를 사용하여 클릭 (클릭 오류 방지)
                driver.execute_script("arguments[0].click();", more_button)
                time.sleep(2)  # 새로운 메뉴 로드를 위한 대기
                
            except Exception as e:
                print("'더보기' 버튼을 더 이상 찾을 수 없거나 클릭할 수 없습니다. 메뉴 로드를 완료합니다.")
                break

        # --- 메뉴 데이터 스크래핑 ---
        print("메뉴 데이터 수집을 시작합니다...")
        menus = driver.find_elements(By.CSS_SELECTOR, 'li.E2jtL')
        if menus:
            for li in menus:
                menu_name, menu_desc, menu_price, menu_recommendation = "", "", "", ""

                # --- 메뉴 이름(span.lPzHi) 수집 ---
                name_elements = li.find_elements(By.CSS_SELECTOR, "span.lPzHi")
                if name_elements:
                    menu_name = name_elements[0].text.strip()

                # --- 메뉴 설명(div.kPogF) 수집 ---
                desc_elements = li.find_elements(By.CSS_SELECTOR, "div.kPogF")
                if desc_elements:
                    menu_desc = desc_elements[0].text.strip()

                recommendation_elements = li.find_elements(By.CSS_SELECTOR, "span.QM_zp span")
                if recommendation_elements:
                    menu_recommendation = recommendation_elements[0].text.strip()

                # --- 메뉴 가격(div.GXS1X) 수집 ---
                # ⭐ em 태그가 있는 경우와 없는 경우 모두 처리
                price_em_elements = li.find_elements(By.CSS_SELECTOR, "div.GXS1X em")
                if price_em_elements:
                    # em 태그가 있으면 그 안의 텍스트를 가격으로
                    menu_price = price_em_elements[0].text.strip()
                else:
                    # em 태그가 없으면 div.GXS1X의 텍스트를 가격으로
                    price_div_elements = li.find_elements(By.CSS_SELECTOR, "div.GXS1X")
                    if price_div_elements:
                        menu_price = price_div_elements[0].text.strip()
                
                # 수집한 정보가 하나라도 있을 경우에만 리스트에 추가
                if menu_name or menu_desc or menu_price or menu_recommendation:
                    menu_sheet.append([menu_name, menu_price, menu_desc, menu_recommendation])


        # 가게 리뷰 페이지 로드
        store_page = 'https://m.place.naver.com/restaurant/' + str(store_id) + '/review/visitor'
        driver.get(store_page)
        time.sleep(3) # 페이지 초기 로드를 위한 짧은 대기
 
        # 📌 2. 세 번째 시트: 'review' (가게 후기)
        review_sheet = xlsx.create_sheet('review') # 'review'라는 이름의 새 시트를 생성합니다.
        review_sheet.append(['content', 'date', 'revisit'])
        
        # --- '더보기' 버튼 클릭 반복 ---
        count = 0
        while True:
            # max_iterations 설정에 따라 반복 중단
            if max_iterations > 0 and count >= max_iterations:
                print(f"지정된 반복 횟수({max_iterations})에 도달하여 '더보기' 클릭을 중단합니다.")
                break

            try:
                # '더보기' 버튼이 클릭 가능할 때까지 최대 10초 대기
                more_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.fvwqf'))
                )
                # JavaScript를 사용하여 클릭 (클릭 오류 방지)
                driver.execute_script("arguments[0].click();", more_button)
                count += 1
                print(f"'{count}'번째 '더보기' 클릭 완료.")
                time.sleep(2)  # 새로운 리뷰 로드를 위한 대기
                
            except Exception as e:
                print("'더보기' 버튼을 더 이상 찾을 수 없거나 클릭할 수 없습니다. 리뷰 로드를 완료합니다.")
                break

        # --- 리뷰 데이터 스크래핑 ---
        print("리뷰 데이터 수집을 시작합니다...")
        reviews = driver.find_elements(By.CSS_SELECTOR, 'li.place_apply_pui.EjjAW')
        if reviews:
            for r in reviews:
                try:
                    content = r.find_element(By.CSS_SELECTOR, 'div.pui__vn15t2').text.strip()
                    date = r.find_element(By.CSS_SELECTOR, 'span.pui__gfuUIT > time').text.strip()
                    
                    revisit_elements = r.find_elements(By.CSS_SELECTOR, 'span.pui__gfuUIT')
                    revisit_text = revisit_elements[1].text.strip() if len(revisit_elements) > 1 else ''

                    revisit_count = ''
                    if "번째 방문" in revisit_text:
                        try:
                            # "번째 방문" 텍스트를 제거하고 정수로 변환
                            revisit_count = int(revisit_text.replace('번째 방문', '').strip())
                        except ValueError:
                            revisit_count = revisit_text # 숫자로 변환 실패 시 원본 텍스트 유지

                    # 엑셀 시트에 데이터 추가
                    review_sheet.append([content, date, revisit_count])
                    
                except Exception as e:
                    print(f"리뷰 하나를 처리하는 중 오류 발생: {e}")
                    # 오류가 발생한 리뷰는 건너뛰고 계속 진행
                    continue

        if xlsx:
            file_name = f'{store_name}_{store_id}_{datetime.datetime.now().strftime("%Y-%m-%d")}.xlsx'
            full_save_path = os.path.join(scraping_file_path, file_name)
            xlsx.save(full_save_path)
            print(f"데이터 수집이 완료되어 '{file_name}' 파일로 저장되었습니다.")

except Exception as e:
    print(f'스크립트 실행 중 오류가 발생했습니다: {e}')

finally:
    # --- 리소스 해제 ---
    if driver:
        driver.quit()
        print("WebDriver를 종료했습니다.")